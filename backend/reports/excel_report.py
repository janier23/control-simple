from openpyxl import Workbook
from openpyxl.styles import Font
import os
from datetime import datetime
from flask import send_file
import io

def generate_excel(data):
    buffer = io.BytesIO()
    buffer.write(b"Reporte PDF")  # ejemplo
    buffer.seek(0)
    os.makedirs("tmp", exist_ok=True)
    filename = f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join("tmp", filename)

    wb = Workbook()

    # HOJA RESUMEN
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "Período"
    ws["B1"] = f"{data['desde']} a {data['hasta']}"
    ws["A3"] = "Ventas"
    ws["B3"] = data["ventas"]
    ws["A4"] = "Gastos"
    ws["B4"] = data["gastos"]
    ws["A5"] = "Ganancia"
    ws["B5"] = data["ganancia"]

    for cell in ["A1", "A3", "A4", "A5"]:
        ws[cell].font = Font(bold=True)

    # HOJA VENTAS
    ws_v = wb.create_sheet("Ventas")
    ws_v.append(["Fecha", "Producto", "Cantidad", "Total", "Usuario"])
    for cell in ws_v[1]:
        cell.font = Font(bold=True)

    for v in data["ventas_detalle"]:
        ws_v.append([
            v["fecha"],
            v["producto"],
            v["cantidad"],
            v["total"],
            v["usuario"]
        ])

    # HOJA GASTOS
    ws_g = wb.create_sheet("Gastos")
    ws_g.append(["Fecha", "Descripción", "Monto", "Usuario"])
    for cell in ws_g[1]:
        cell.font = Font(bold=True)

    for g in data["gastos_detalle"]:
        ws_g.append([
            g["fecha"],
            g["descripcion"],
            g["monto"],
            g["usuario"]
        ])

    wb.save(path)
    return path
